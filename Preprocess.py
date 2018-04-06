import os
import pandas as pd
import nltk
import string
import cPickle
import numpy as np
from nltk.corpus import stopwords
from gensim.models import Word2Vec
from openpyxl.reader.excel import load_workbook
from datetime import datetime, timedelta
from sklearn.preprocessing import MinMaxScaler



def word_token(sentence, vocab = None, stopwords_list = None):
    words = nltk.word_tokenize(sentence)
    if stopwords_list is not None:
        words = list(filter(lambda x: x not in stopwords_list, words))
    if vocab is not None:
        words = list(filter(lambda x: x in vocab, words))
    return words

def get_w2v_model(model_name, news_dim, min_count = 5, content_file = None):
    if os.path.exists(model_name):
        print 'Loading word2vector model from', model_name
        model = Word2Vec.load(model_name)
        return model

    if content_file is None:
        print 'Please input the name of the content file to build word2vec model.'
        return

    if not os.path.exists(content_file):
        print 'content file %s doesn\'t exist. Please correct and try again.' % content_file
        return
    with open(content_file, 'r') as f:
        train_texts = f.read().decode('utf-8')
        train_texts = filter(lambda x: len(x.encode('utf-8'))==1, train_texts)  # filter strange unicode

    stopwords_list = list(string.punctuation) + stopwords.words('english')
    # tokenize each news
    sents = nltk.sent_tokenize(train_texts)
    sentences = list(map(lambda s: word_token(s, None, stopwords_list), sents))

    # build word2vec model
    model = Word2Vec(sentences, min_count=min_count, size=news_dim)
    model.save(model_name)
    print 'Building word2vector model, and save it into', model_name
    return model

def news2vector(news_list, wv):
    '''
    :param news_list: list of news tuple (time, title+content)
    :param wv: word2vector model (model.wv)
    :param all_words: words in vocabulary
    :return:
    '''

    all_words = wv.vocab.keys()

    # tranfer sentences into  words
    words_list = list(map(lambda news: (news[0], word_token(news[1], all_words)), news_list)) # (time, list_of_words)

    # filter empty list
    words_list = list(filter(lambda wl: len(wl[1]) != 0, words_list))

    # word2vector
    vector_list = list(map(lambda words: (words[0], wv[words[1]]), words_list)) # (time, list_of_vectors)

    # # evenly sum all the word-vector in one sentence
    # vector_list = list(map(lambda v: v.sum(axis=0), vector_list))

    return vector_list


def news_process(news_list, w2v_model, target_news_file):
    # news to vectors, with format (time, list_of_vectors)
    vector_list = news2vector(news_list, w2v_model.wv)

    # get the max number of words for each news
    max_num_words = max(list(map(lambda x: len(x[1]), vector_list)))
    # padding news
    time_list = list(map(lambda x: x[0], vector_list))
    news_matrix = np.array(list(map(lambda x: np.pad(x[1], ((0, max_num_words-x[1].shape[0]) , (0, 0)), 'constant', constant_values=(0, 0)), vector_list)))

    # save news info into files
    print 'save news information into %s' % target_news_file
    with open(target_news_file, 'w') as f:
        cPickle.dump((news_matrix, time_list), f)

    return news_matrix, time_list

def news_loader(filename, w2v_model, target_news_file):
    if target_news_file is not None and os.path.exists(target_news_file):
        print 'Load news data from ', target_news_file
        with open(target_news_file, 'r') as f:
            news_matrix, time_list = cPickle.load(f)
        return news_matrix, time_list

    dir = filename[:filename.find('/')+1]
    target_news_file = dir + 'news_info.pkl'
    if not os.path.exists(filename):
        print '%s doesn\'t exist. Please correct the filename and try again.' % filename
        return

    wb = load_workbook(filename)
    ws = wb.active
    row = ws.max_row
    news_list = list()
    for i in range(row, 1, -1):
        title = ws['A'+str(i)].value
        released_time = ws['B'+str(i)].value
        content = ws['D'+str(i)].value
        dt = datetime.strptime(released_time[:19], '%Y-%m-%d %H:%M:%S')
        news_list.append((dt, title + '. ' + content))

    news_matrix, time_list = news_process(news_list, w2v_model, target_news_file)
    return  news_matrix, time_list

def label(x, up, down):
    if x <= down:
        return 0
    elif x < up:
        return 1
    else:
        return 2

def price_label(all_flucts, target_time_fluct):
    '''

    :param all_flucts:
    :param target_time_fluct:
    :return:
    '''
    columns = target_time_fluct.columns
    num = len(all_flucts)
    labels = np.zeros((len(target_time_fluct), len(columns)), dtype=np.int64)
    for i, c in enumerate(columns):
        tmp_series = all_flucts[c].copy().as_matrix()
        tmp_series.sort()
        down = tmp_series[num/3]
        up = tmp_series[num*2/3]
        cur_label_list = list(map(lambda x: label(x, up, down), target_time_fluct[c]))
        cur_label_list = np.array(cur_label_list).reshape(-1,)
        labels[:, i] = cur_label_list
    labels = pd.DataFrame(labels, index=target_time_fluct.index, columns=columns)
    return labels


def price_process(price, time_list, time_gaps, all_price_fluc_file, target_price_fluc_file):
    '''
    process price to get price_features and (up, stay, down)labels
    :param price:
    :param time_list:
    :param all_price_fluc_file:
    :param target_price_fluc_file:
    :return:
    '''
    # process price
    weighted_price = price['Weighted_Price'].copy()
    price.drop('Weighted_Price', axis=1, inplace=True)

    # check if related file exists
    if all_price_fluc_file is not None and target_price_fluc_file is not None and os.path.exists(all_price_fluc_file) and os.path.exists(target_price_fluc_file):
        print 'Load price information from %s and %s.' % (all_price_fluc_file, target_price_fluc_file)
        with open(all_price_fluc_file, 'r') as f:
            all_flucts = cPickle.load(f)
        with open(target_price_fluc_file, 'r') as f:
            price_features, target_time_fluct = cPickle.load(f)
    else:
        all_price_fluc_file, target_price_fluc_file = 'all_flucts.pkl', 'price_info.pkl'

        time_intervals = list()
        for t in time_gaps:
            if t.endswith('h'):
                time_intervals.append(timedelta(hours=int(t[:-1])))
            elif t.endswith('d'):
                time_intervals.append(timedelta(days=int(t[:-1])))
            elif t.endswith('w'):
                time_intervals.append(timedelta(weeks=int(t[:-1])))
            else:
                print 'time gap has incorrect time interval, please check.'
                return

        all_time_points = price.index

        all_flucts = pd.DataFrame(index=all_time_points, columns=time_gaps)
        target_time_fluct = pd.DataFrame(index=time_list, columns=time_gaps)

        price_features = pd.DataFrame(index=time_list, columns=price.columns)

        for dt in all_time_points:
            flag = 0 # to identify whether current dt in time_list
            # get features
            if dt in time_list:
                for c in price.columns:
                    price_features.loc[dt, c] = price.loc[dt, c]
                flag = 1
            # get fluctuations for all time point
            original_price = weighted_price[dt]
            for tg, time_diff in zip(time_gaps, time_intervals):
                target_time = dt + time_diff
                if target_time in all_time_points:
                    fluct = (weighted_price.loc[target_time] - original_price) / original_price
                    all_flucts.loc[dt, tg] = fluct
                    if flag:
                        target_time_fluct.loc[dt, tg] = fluct
        # save all_fluct and labels into pkl files
        with open(all_price_fluc_file,'w') as f:
            cPickle.dump(all_flucts, f)

        with open(target_price_fluc_file, 'w') as f:
            cPickle.dump((price_features, target_time_fluct), f)

    # scale the features
    sc = MinMaxScaler()
    sc.fit(price.values)
    price_features = sc.transform(price_features.values)
    # relabel the price_change
    labels = price_label(all_flucts, target_time_fluct)

    return price_features, labels


def price_loader(price_file, time_list, time_gaps, all_price_fluc_file, target_price_fluc_file):
    if not os.path.exists(price_file):
        print '%s doesn\'t exist. Please correct the filename and try again.' % price_file
        return

    # read time-price information from price file
    df = pd.read_csv(price_file)
    # generate 'date' from 'timestamp'
    df['date'] = pd.to_datetime(df['Timestamp'], unit='s')
    # filter data before 2017, and set 'date' as index
    df = df[df['date'] > '2017'].set_index('date')
    # delete 'Timestamp' column
    df.drop('Timestamp', axis=1, inplace=True)

    price_feature, labels = price_process(df, time_list, time_gaps, all_price_fluc_file, target_price_fluc_file)

    return price_feature, labels
